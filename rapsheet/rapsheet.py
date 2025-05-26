#!/usr/bin/env python3
import sys
import glob
import pandas as pd
import shutil
import os
from pathlib import Path
import argparse
import re
import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def fix_bom_header(file_name):
    with open(file_name, "rb") as f:
        content = f.read()
        # Check if BOM is already present
        if not content.startswith(b'\xef\xbb\xbf'):
            with open(file_name, "wb") as f_out:
                f_out.write(b'\xef\xbb\xbf' + content)

def extract_inv_number(file_name):
    match = re.search(r'INV(\d+)', file_name)
    return match.group(0) if match else None

def check_consistent_inv_numbers(files):
    inv_numbers = {extract_inv_number(file) for file in files if extract_inv_number(file)}
    if len(inv_numbers) > 1:
        return False, "Conflict in the Incident number file names, check your input data"
    return True, inv_numbers.pop() if inv_numbers else None

def generate_sheet_name(file_name):
    """Generate a clean sheet name by removing INV/SIR numbers and replacing underscores."""
    base_name = Path(file_name).stem
    cleaned_name = re.sub(r'(INV|SIR)(\d+)', '', base_name).replace('_', ' ').strip()
    return cleaned_name[:31]  # Excel sheet names are limited to 31 characters.

def autofit_columns(worksheet):
    """Automatically adjust column widths based on maximum cell length, optimized."""
    for col in worksheet.columns:
        column = col[0].column_letter
        
        # Collect all cell values, but strip whitespace and filter out pure empty strings
        cell_values = [str(cell.value).strip() for cell in col if cell.value not in [None, ""]]
        
        # Remove empty strings, whitespace-only, and newline-only cells
        cell_values = [value for value in cell_values if value and not value.isspace()]

        # Calculate maximum length only if there are valid values
        if cell_values:
            max_length = max(len(value) for value in cell_values)
            adjusted_width = min(max(max_length + 2, 10), 80)
        else:
            adjusted_width = 5  # For completely empty columns, set minimal width
        
        # Apply width setting
        worksheet.column_dimensions[column].width = adjusted_width



def apply_text_wrapping(worksheet):
    """Apply text wrapping to cells with content longer than 80 characters, optimized."""
    wrap_alignment = Alignment(wrapText=True)
    
    # Use generator expression for memory efficiency
    cells_to_wrap = (cell for row in worksheet.iter_rows() for cell in row 
                     if cell.value and len(str(cell.value).strip()) > 80)
    
    # Apply wrapping in one sweep
    for cell in cells_to_wrap:
        cell.alignment = wrap_alignment


def convert_csv_to_xlsx(input_files, output_dir, template_path=None):
    consistent, inv_number = check_consistent_inv_numbers(input_files)
    if not consistent:
        inv_number = "Incident_Report"

    output_file = os.path.join(output_dir, f"{inv_number}_Report_{datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx")

    template_exists = template_path and os.path.exists(template_path)
    if template_path and not template_exists:
        print(f"Warning: Specified template '{template_path}' not found. Creating new Excel file.")

    if template_exists:
        shutil.copy(template_path, output_file)

    mode = 'a' if os.path.exists(output_file) else 'w'
    with pd.ExcelWriter(output_file, engine='openpyxl', mode=mode) as writer:
        existing_sheets = writer.book.sheetnames if mode == 'a' else []
        for file_name in input_files:
            if not os.path.exists(file_name):
                print(f"File not found, skipping: {file_name}")
                continue

            fix_bom_header(file_name)
            base_name = generate_sheet_name(file_name)

            if base_name in existing_sheets:
                print(f"Sheet '{base_name}' already exists, skipping: {file_name}")
                continue

            try:
                df = pd.read_csv(file_name, low_memory=False)
            except Exception as e:
                print(f"Error reading CSV file '{file_name}': {e}")
                continue

            df.to_excel(writer, sheet_name=base_name, index=False)
            worksheet = writer.sheets[base_name]
            worksheet.freeze_panes = 'A2'
            autofit_columns(worksheet)
            apply_text_wrapping(worksheet)

            print(f"Processed file: '{file_name}'")

    print(f"Report saved to: {output_file}")

   
def main():
    parser = argparse.ArgumentParser(description="CSV to Excel Report Generator")
    parser.add_argument("-f", "--files", nargs="+", help="List of CSV files to process", default=[])
    parser.add_argument("-d", "--directory", help="Directory containing CSV files", default="")
    parser.add_argument("-o", "--output", help="Output directory for the report", required=True)
    parser.add_argument("-t", "--template", help="Path to the template Excel file", required=False)
    parser.add_argument("-u", "--update", action='store_true', help="Update using the Excel template in the directory")

    args = parser.parse_args()

    input_files = args.files
    template_file = args.template

    # Check for directory and handle the -u update logic
    if args.directory:
        directory_files = glob.glob(os.path.join(args.directory, "*.csv"))
        input_files.extend(directory_files)
        if args.update:
            # Update template handling for directory
            xlsx_files = glob.glob(os.path.join(args.directory, "*.xlsx"))
            if len(xlsx_files) == 1:
                template_file = xlsx_files[0]
            elif len(xlsx_files) > 1:
                print("Error: More than one .xlsx file found in the directory.")
                sys.exit(1)

    if args.update and not template_file:
        # Handle -u update for file list if no template has been set yet
        potential_templates = [f for f in input_files if f.lower().endswith('.xlsx')]
        if len(potential_templates) == 1:
            template_file = potential_templates[0]
            input_files = [f for f in input_files if f != template_file]  # Remove the template from input files
        elif len(potential_templates) > 1:
            print("Error: More than one .xlsx file found in the input file list.")
            sys.exit(1)

    # Validate input files
    input_files = [f for f in input_files if f.lower().endswith('.csv')]

    if not input_files:
        print("Error: No CSV files specified. Use -f for file(s) or -d for a directory.")
        sys.exit(1)

    convert_csv_to_xlsx(input_files, args.output, template_file)

if __name__ == "__main__":
    main()

